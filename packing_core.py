"""
packing_core.py
===============
Core logic for turning a parcel master workbook into a "<Metal> Packing List".

Designed to run both locally and on Streamlit Community Cloud: every function
works on in-memory bytes (no disk paths required) and returns bytes, so it can
be wired straight into a file uploader / download button.

Public API
----------
    list_sheets(file_bytes)               -> [sheet names]
    suggest_purity_unit(name)             -> 'kt' | 'fineness' | 'none'
    generate_packing_list(file_bytes, sheet, metal=None,
                          purity_unit=None, markup=0.04)
                                          -> (output_bytes, info_dict)
"""

import io
import os
import re
import shutil
import tempfile

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage


# ----------------------------- small helpers --------------------------------
def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s is not None else ""


def num(v):
    return v if isinstance(v, (int, float)) else 0


def clean(v):
    s = "" if v is None else str(v).strip()
    return "" if s.lower() in ("none", "0") else s


# ------------------------ source-sheet introspection ------------------------
def _find_header_row(ws, max_scan=8):
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(r, c).value) == "sno":
                return r
    raise ValueError("Could not find the header row (no 'S.NO' cell found).")


def _header_index(ws, hrow):
    return [(c, norm(ws.cell(hrow, c).value), ws.cell(hrow, c).value)
            for c in range(1, ws.max_column + 1)]


def _col_by(cols, *names, occurrence=1):
    targets = {norm(n) for n in names}
    seen = 0
    for c, nh, _ in cols:
        if nh in targets:
            seen += 1
            if seen == occurrence:
                return c
    return None


def _section_cols(cols, start, end):
    ctwt = total = None
    for c, nh, _ in cols:
        if c < start or c > end:
            continue
        if ctwt is None and nh.startswith("ct"):
            ctwt = c
        if total is None and nh.endswith("total"):
            total = c
    return ctwt, total


# --------------------------------- public -----------------------------------
def list_sheets(file_bytes):
    """Return the sheet names of an uploaded .xlsx (read-only, fast)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def suggest_purity_unit(name):
    n = (name or "").lower()
    if "gold" in n:
        return "kt"
    if "silver" in n:
        return "fineness"
    return "none"


def generate_packing_list(file_bytes, sheet, metal=None, purity_unit=None, markup=0.04):
    """
    Build a packing-list workbook from `sheet` inside the uploaded file.

    Returns (output_bytes, info) where info = {sheet, metal, items, purity_unit,
    markup, output_name}.
    """
    metal = metal or sheet
    purity_unit = purity_unit or suggest_purity_unit(metal)

    wb = load_workbook(io.BytesIO(file_bytes))                 # formulas + images
    wbv = load_workbook(io.BytesIO(file_bytes), data_only=True)  # cached values
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    sv = wbv[sheet]

    # keep only the chosen source sheet in the output workbook
    for s in list(wb.sheetnames):
        if s != sheet:
            del wb[s]
    src = wb[sheet]

    # ---- locate columns by header --------------------------------------
    hrow = _find_header_row(sv)
    cols = _header_index(sv, hrow)
    data_start = hrow + 1

    c_sno   = _col_by(cols, "S.NO")
    c_tag   = _col_by(cols, "Tag No")
    c_col   = _col_by(cols, "COLOUR", "Colour")
    c_item  = _col_by(cols, "ITEAM", "ITEM", "Item")
    c_pcs   = _col_by(cols, "PCS/PAIR")
    c_gross = _col_by(cols, "G Wt")
    c_net   = _col_by(cols, "N Wt")
    c_pur   = _col_by(cols, "Declared Purity")
    c_fine  = _col_by(cols, "Total", occurrence=1)
    c_metal = _col_by(cols, "Metal Amount")
    c_val   = _col_by(cols, "TOTAL Labour Amount")

    # the fine-metal 'Total' sits before Metal Amount; if the only 'Total' is the
    # grand total at the far right (e.g. Brass has no fine-metal column), drop it
    if c_fine and c_metal and c_fine > c_metal:
        c_fine = None

    if None in (c_sno, c_tag, c_gross):
        raise ValueError("Source sheet is missing essential columns (S.NO / Tag No / G Wt).")

    fine_label = "Total Fine Metal"
    for _, _, raw in cols:
        if raw and "ratein" in norm(raw):
            m = re.search(r"(\d{3})", str(raw))
            if m:
                fine_label = f"Total Fine Metal {m.group(1)}"

    cz = _col_by(cols, "CZ Classification")
    st = _col_by(cols, "Stone Classification")
    pl = _col_by(cols, "Pearls")
    di = _col_by(cols, "Diamond")
    end = sv.max_column
    anchors = sorted(a for a in (cz, st, pl, di) if a)

    def bound(a):
        nxt = [x for x in anchors if x > a]
        return (nxt[0] - 1) if nxt else end

    cz_ct, cz_tot = _section_cols(cols, cz, bound(cz)) if cz else (None, None)
    st_ct, st_tot = _section_cols(cols, st, bound(st)) if st else (None, None)
    pl_ct, pl_tot = _section_cols(cols, pl, bound(pl)) if pl else (None, None)
    di_ct, di_tot = _section_cols(cols, di, bound(di)) if di else (None, None)
    c_st_type = (st + 1) if st else None

    def gv(r, c):
        return sv.cell(r, c).value if c else None

    def purity_str(p):
        if purity_unit == "kt":
            return f"{round((p or 0) * 24)}Kt"
        if purity_unit == "fineness":
            return f"{round((p or 0) * 1000)}"
        return ""

    def describe(p, colour, item):
        col = clean(colour)
        if col.lower() == metal.lower():   # avoid "Silver Silver" / "Brass Brass"
            col = ""
        parts = [purity_str(p), col, metal, clean(item)]
        return " ".join(x for x in parts if str(x).strip())

    def stone_type(r):
        parts = []
        if cz_ct and num(gv(r, cz_ct)) > 0:
            parts.append(clean(gv(r, cz)))
        if st_ct and num(gv(r, st_ct)) > 0:
            parts.append((clean(gv(r, st)) + " " + clean(gv(r, c_st_type))).strip())
        if pl_ct and num(gv(r, pl_ct)) > 0:
            parts.append("Pearls")
        if di_ct and num(gv(r, di_ct)) > 0:
            parts.append(clean(gv(r, di)))
        return ", ".join(x for x in parts if x)

    last = hrow
    for r in range(data_start, sv.max_row + 1):
        if gv(r, c_sno) is not None or gv(r, c_tag) is not None:
            last = r

    # extract source images and repair their file handles for re-save
    tmpdir = tempfile.mkdtemp(prefix="pkimg_")
    try:
        imgmap = {}
        for im in src._images:
            rr = im.anchor._from.row + 1
            data = im._data()
            path = os.path.join(tmpdir, f"r{rr}.png")
            with open(path, "wb") as fh:
                fh.write(data)
            imgmap[rr] = (path, im.width, im.height)
            im.ref = io.BytesIO(data)

        recs = []
        for r in range(data_start, last + 1):
            recs.append(dict(
                sno=gv(r, c_sno), tag=gv(r, c_tag),
                item=describe(gv(r, c_pur), gv(r, c_col), gv(r, c_item)),
                colour=gv(r, c_col), purity=gv(r, c_pur), img=imgmap.get(r),
                pcs=gv(r, c_pcs), gross=gv(r, c_gross), net=gv(r, c_net),
                carat=sum(num(gv(r, c)) for c in (cz_ct, st_ct, pl_ct, di_ct) if c),
                fine=gv(r, c_fine), metal=gv(r, c_metal), valadd=gv(r, c_val),
                stype=stone_type(r),
                agg=sum(num(gv(r, c)) for c in (cz_tot, st_tot, pl_tot, di_tot) if c),
            ))

        if c_pur:
            recs.sort(key=lambda x: (-(x["purity"] or 0), str(x["colour"] or "")))
        else:
            recs.sort(key=lambda x: str(x["colour"] or ""))

        _write_sheet(wb, metal, fine_label, recs, markup)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    info = dict(sheet=sheet, metal=metal, items=len(recs),
               purity_unit=purity_unit, markup=markup,
               output_name=f"{metal} Packing List.xlsx")
    return out.getvalue(), info


# --------------------------- sheet construction -----------------------------
def _write_sheet(wb, metal, fine_label, recs, markup):
    ws = wb.create_sheet(f"{metal} Packing List")
    hcolour = "1F4E78" if "gold" in metal.lower() else "595959"
    hfont = Font(name="Arial", bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color=hcolour)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    infill = PatternFill("solid", start_color="FFF2CC")
    inblue = Font(name="Arial", color="0000FF", bold=True)
    lab = Font(name="Arial", bold=True)

    ws["A1"] = f"{metal} Rate (INR)"; ws["A1"].font = lab
    ws["B1"].fill = infill; ws["B1"].font = inblue; ws["B1"].border = border; ws["B1"].number_format = "#,##0.00"
    ws["D1"] = "USD to INR"; ws["D1"].font = lab
    ws["E1"].fill = infill; ws["E1"].font = inblue; ws["E1"].border = border; ws["E1"].number_format = "#,##0.0000"

    headers = ["S.NO", "Tag No", "Item", "Colour", "Declared Purity", "Picture",
               "PCS/PAIR", "Gross Wt Gms", "Stone / Diamond Carat Wt", "Net Wt",
               fine_label, "Metal Amount in INR", "Metal Amount in USD",
               "Value addition in INR", "Value Addition in USD", "Stone Type",
               "Aggregated Stone Price", "Mark up %", "Marked Up Stone Price",
               "Stone Value (USD)", "Total Amount in USD"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font = hfont; cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    MAX = 90
    for i, rec in enumerate(recs):
        out = 3 + i
        ws.cell(out, 1, i + 1); ws.cell(out, 2, rec["tag"]); ws.cell(out, 3, rec["item"])
        ws.cell(out, 4, rec["colour"]); ws.cell(out, 5, rec["purity"])
        ws.cell(out, 7, rec["pcs"]); ws.cell(out, 8, rec["gross"]); ws.cell(out, 9, rec["carat"])
        ws.cell(out, 10, rec["net"]); ws.cell(out, 11, rec["fine"]); ws.cell(out, 12, rec["metal"])
        ws.cell(out, 13, f'=IF($E$1="","",L{out}/$E$1)')
        ws.cell(out, 14, rec["valadd"])
        ws.cell(out, 15, f'=IF($E$1="","",N{out}/$E$1)')
        ws.cell(out, 16, rec["stype"]); ws.cell(out, 17, rec["agg"]); ws.cell(out, 18, markup)
        ws.cell(out, 19, f"=Q{out}*(1+R{out})")
        ws.cell(out, 20, f'=IF($E$1="","",S{out}/$E$1)')
        ws.cell(out, 21, f'=IF($E$1="","",M{out}+O{out}+T{out})')
        for c in range(1, 22):
            cell = ws.cell(out, c)
            cell.font = Font(name="Arial"); cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c == 6 else "left",
                                       wrap_text=(c in (3, 16)))
        ws.cell(out, 5).number_format = "0.000"
        for c in (8, 9, 10, 11):
            ws.cell(out, c).number_format = "#,##0.000"
        for c in (12, 13, 14, 15, 17, 19, 20, 21):
            ws.cell(out, c).number_format = "#,##0.00"
        ws.cell(out, 18).number_format = "0%"; ws.cell(out, 18).font = inblue
        if rec["img"]:
            path, w, h = rec["img"]
            img = XLImage(path)
            scale = min(MAX / w, MAX / h, 1.0)
            img.width = w * scale; img.height = h * scale
            ws.add_image(img, f"F{out}")
            ws.row_dimensions[out].height = img.height * 0.75 + 6

    widths = {"A": 6, "B": 14, "C": 26, "D": 10, "E": 12, "F": 14, "G": 10,
              "H": 12, "I": 14, "J": 10, "K": 14, "L": 15, "M": 15, "N": 15,
              "O": 15, "P": 28, "Q": 16, "R": 9, "S": 16, "T": 15, "U": 15}
    for col, wd in widths.items():
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A3"

    # packing list first, source sheet second
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)
    return ws
